#!/usr/bin/python3
import os
import sys
import re
import openpyxl
argv_l = sys.argv
excel_name = argv_l[1]
#  excel_name= "RVSEED.xlsx"
input_file= os.getcwd()+"/"+excel_name
excel_name= re.sub(r'\.\w+','',excel_name)
excel_name= re.sub(r'\_\w+','',excel_name)

#生成一个test.v，其中写入的是将test.xlsx中的信息转化为verilog的代码框架
gen_file= os.getcwd()+"/../rtl/"+excel_name.lower()+"/"+"regs_"+excel_name.lower()+'.v'
print(gen_file)
#每次运行的时候检测test.v是否存在，如果存在则删除
if os.path.exists(gen_file):
    os.remove(gen_file)
f=open(gen_file,"a")
#1、打开excel的工作表
wb = openpyxl.load_workbook(input_file)
#2、获取工作表的名字test,返回的是一个list
sheet_list= wb.sheetnames
sheet_num= len(sheet_list)
for j in range(sheet_num):
    sheet_name=''.join(sheet_list[j])
    f.write("//--------------------------------------------------------------------------------" + '\n')
    f.write("// " + sheet_name + '\n')
    f.write("//--------------------------------------------------------------------------------" + '\n')
    f.write("module REGS_"+ str(sheet_name).upper() +"(" + '\n')
    f.write(str("    input  wire                            clk_reg,").ljust(40)+ '\n')
    f.write(str("    input  wire                            rst_reg_n,").ljust(40)+ '\n')

    f.write(str("    input  wire                            reg_wen,    // register write enable").ljust(40)+ '\n')
    f.write(str("    input  wire [`REG_ADDR_WIDTH   -1:0]   reg_waddr,  // register write address").ljust(40)+ '\n')
    f.write(str("    input  wire [`REG_DATA_WIDTH   -1:0]   reg_wdata,  // register write data").ljust(40)+ '\n')
    f.write('\n')    
    f.write(str("    input  wire [`REG_ADDR_WIDTH   -1:0]   reg1_raddr, // register 1 read address").ljust(40)+ '\n')
    f.write(str("    input  wire [`REG_ADDR_WIDTH   -1:0]   reg2_raddr, // register 2 read address").ljust(40)+ '\n')
    f.write(str("    output reg  [`REG_DATA_WIDTH   -1:0]   reg1_rdata, // register 1 read data").ljust(40)+ '\n')
    f.write(str("    output reg  [`REG_DATA_WIDTH   -1:0]   reg2_rdata  // register 2 read data").ljust(40)+ '\n')
    f.write(");"+ '\n')
    table= wb[sheet_name] #open sheet by name
    row_num = table.max_row
    col_num = table.max_column
    name    = []
    offset  = []
    suffix  = []
    attr    = []
    default = []
    comment = []
    for i in range(row_num):
        name.append(table.cell(row=i+1,column=1).value) #excel col1
        offset.append(table.cell(row=i+1,column=2).value) #excel col2
        suffix.append(table.cell(row=i+1,column=2).value) #excel col2
        attr.append(table.cell(row=i+1,column=3).value) #excel col3
        default.append(table.cell(row=i+1,column=4).value) #excel col4
        comment.append(table.cell(row=i+1,column=5).value) #excel col4
        #  default=str(default)
    
    for i in range(row_num-1): # pre replace 
        offset[i+1]=str(offset[i+1]).replace("0x","16'h")
        suffix[i+1]=str(suffix[i+1]).replace("0x","h").lower()
        default[i+1]=str(default[i+1]).replace("0x","32'h")


    for i in range(row_num-1):
        f.write("// " + str(name[i+1]+" ").lower()+"("+ str(comment[i+1]).lower()+")"+ '\n')
        if str(attr[i+1])=="RW":
            f.write(str("wire wen_"+suffix[i+1]+" = reg_wen & (reg_waddr == "+offset[i+1]+");").lower() + '\n') 
        else:
            f.write(str("wire wen_"+suffix[i+1]+" = 1'b0;") + '\n')
        f.write(str("wire [`REG_DATA_WIDTH-1:0] rdata_"+suffix[i+1]+";") + '\n')
        f.write(str("RW_REG #(32,"+default[i+1])+") U_"+str(name[i+1]).upper()+"_"+str(suffix[i+1]).upper()+" ("+'\n')
        f.write(str("    .clk_reg").ljust(20)+str("(clk_reg").ljust(20)+str("),") +'\n')
        f.write(str("    .rst_reg_n").ljust(20)+str("(rst_reg_n").ljust(20)+str("),") +'\n')
        f.write(str("    .wen").ljust(20)+str("(wen_"+suffix[i+1]).ljust(20)+str("),") +'\n')
        f.write(str("    .data_in").ljust(20)+str("(reg_wdata").ljust(20)+str("),") +'\n')
        f.write(str("    .data_out").ljust(20)+str("(rdata_"+suffix[i+1]).ljust(20)+str(")") +'\n')
        f.write(str(");"+'\n'))
        f.write('\n')


    f.write(str("always @(*) begin")+ '\n')
    f.write(str("    case(reg1_raddr)")+ '\n')
    for i in range(row_num-1):
        f.write(str("        "+offset[i+1]+" : reg1_rdata = "+str("rdata_"+suffix[i+1]).ljust(10)+str(";") +'\n'))
    f.write(str("        default : reg1_rdata = 32'h0;")+ '\n')
    f.write(str("    endcase")+ '\n')
    f.write(str("end")+ '\n')

    f.write(str("always @(*) begin")+ '\n')
    f.write(str("    case(reg2_raddr)")+ '\n')
    for i in range(row_num-1):
        f.write(str("        "+offset[i+1]+" : reg2_rdata = "+str("rdata_"+suffix[i+1]).ljust(10)+str(";") +'\n'))
    f.write(str("        default : reg2_rdata = 32'h0;")+ '\n')
    f.write(str("    endcase")+ '\n')
    f.write(str("end")+ '\n')


    f.write('\n')
    f.write("endmodule"+'\n')
                
f.close()
