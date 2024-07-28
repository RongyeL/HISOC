#!/usr/bin/python3
import os
import sys
import re
import openpyxl
argv_l = sys.argv
excel_name = argv_l[1]
#  excel_name= "RVSEED.xlsx"
input_file= os.getcwd()+"/"+excel_name
excel_name= re.sub(r'\.\w+','',excel_name)
excel_name= re.sub(r'\_\w+','',excel_name)

#生成一个test.v，其中写入的是将test.xlsx中的信息转化为verilog的代码框架
gen_file= os.getcwd()+"/"+excel_name.lower()+'_define.v'
#每次运行的时候检测test.v是否存在，如果存在则删除
if os.path.exists(gen_file):
    os.remove(gen_file)
f=open(gen_file,"a")
#1、打开excel的工作表
wb = openpyxl.load_workbook(input_file)
#2、获取工作表的名字test,返回的是一个list
sheet_list= wb.sheetnames
sheet_num= len(sheet_list)
for j in range(sheet_num):
    sheet_name=''.join(sheet_list[j])
    f.write("//--------------------------------------------------------------------------------" + '\n')
    f.write("// " + sheet_name + '\n')
    f.write("//--------------------------------------------------------------------------------" + '\n')
    table= wb[sheet_name] #open sheet by name
    row_num = table.max_row
    col_num = table.max_column
    field=[]
    width=[]
    for i in range(row_num):
        field.append(table.cell(row=i+1,column=1).value) #excel col1
        width.append(table.cell(row=i+1,column=2).value) #excel col2

    
    f.write("// Width define" + '\n')
    for i in range(row_num-1):
        f.write("`define " + sheet_name+ "_" + str(field[i+1] + "_W").ljust(31).upper() + str(width[i+1]) + '\n')

    f.write('\n')
    f.write("// Position define" + '\n')
    for i in range(row_num-1):
        if(i==0):
            f.write("`define " + sheet_name+ "_" + str(field[i+1] + "_LSB").ljust(30).upper() + " 0" + '\n')
        else:
            f.write("`define " + sheet_name+ "_" + str(field[i+1] + "_LSB").ljust(30).upper() + " `" + sheet_name+ "_" + str(field[i] + "_MSB").upper() + '\n')

        f.write("`define " + sheet_name+ "_" + str(field[i+1] + "_MSB").ljust(30).upper() + "(`" + sheet_name+ "_" + str(field[i+1] + "_W-1)").upper().ljust(15) + "+" + "`" + sheet_name+ "_" + str(field[i+1] + "_LSB").upper() + '\n')
        f.write("`define " + sheet_name+ "_" + str(field[i+1]).ljust(30).upper() + " `" + sheet_name+ "_" + str(field[i+1] + "_MSB").upper().ljust(15) + ":`" + sheet_name+ "_" + str(field[i+1] + "_LSB").upper() + '\n')
        f.write('\n')
            
f.close()
